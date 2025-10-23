import tempfile
import pdfplumber
import re
import os

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ReadPdfPassport(APIView):
    parser_classes = (MultiPartParser, FormParser)
    
    def post(self, request):
        if 'file' not in request.FILES:
            return Response(
                {"error": "Файл не найден в запросе"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        pdf_file = request.FILES['file']
        
        try:
            # Сохраняем файл временно
            with open('temp.pdf', 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)
            
            # Извлекаем заголовок и адрес
            passport_info = self.extract_passport_info('temp.pdf')
            
            # Извлекаем таблицу направлений
            table_data = self.extract_directions_table('temp.pdf')
            
            return Response({
                "success": True,
                "passport_info": passport_info,
                "table_data": table_data,
                "total_directions": len(table_data)
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {"error": f"Ошибка обработки файла: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        finally:
            # Удаляем временный файл
            import os
            if os.path.exists('temp.pdf'):
                os.remove('temp.pdf')
    
    def extract_passport_info(self, pdf_path):
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = pdf.pages[0].extract_text()
                if not text:
                    return {"number": "", "address": ""}

                # Ищем номер паспорта
                number_match = re.search(r'Паспорт[^№]*№\s*(\d+)', text)
                passport_number = number_match.group(1) if number_match else ""
                
                # Ищем адрес после номера
                address = ""
                if number_match:
                    # Берем текст после номера до конца строки или до ключевых слов
                    rest_of_text = text[number_match.end():]
                    address_match = re.search(r'[-–—]?\s*([^\n]+?)(?:\n|$|Паспорт|светофорного)', rest_of_text)
                    if address_match:
                        address = address_match.group(1).strip()
                
                return {"number": passport_number, "address": address}
                
        except Exception as e:
            print(f"Ошибка извлечения: {e}")
            return {"number": "", "address": ""}

    def extract_directions_table(self, pdf_path):
        """Извлечение таблицы направлений из PDF по ключевым заголовкам"""
        table_data = []

        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""

            # Собираем весь текст из PDF
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    full_text += page_text + "\n"

            # Ключевые заголовки для поиска таблицы
            key_headers = [
                '№ нап.',
                'Тип направления', 
                'Фазы, в кот. участ. направ.',
                'Светофоры',
                '"Запрет"',
                '"Разрешение"'
            ]

            lines = full_text.split('\n')
            table_start_index = -1

            # Ищем начало таблицы по ключевым заголовкам
            for i, line in enumerate(lines):
                # Проверяем, содержит ли строка хотя бы 2 ключевых заголовка
                found_headers = sum(1 for header in key_headers if header in line)
                if found_headers >= 2:
                    table_start_index = i
                    print(f"Найдена таблица на строке {i}: {line}")
                    break

            if table_start_index == -1:
                print("Таблица не найдена по ключевым заголовкам")
                return []

            # Пропускаем строку с заголовками и ищем данные
            data_lines = []
            for line in lines[table_start_index + 1:]:
                line = line.strip()
                if not line:
                    continue

                # Ищем строки, которые начинаются с цифры (номера направлений)
                if re.match(r'^\s*\d+\s+', line):
                    data_lines.append(line)
                    print(f"Найдена строка с данными: {line}")
                # Если нашли новую секцию - заканчиваем
                elif line.startswith('___') or line.startswith('**') or 'страница' in line.lower() or 'Информационные секции' in line:
                    break

            # Парсим найденные строки
            for line in data_lines:
                direction_data = self.parse_direction_line(line)
                if direction_data:
                    table_data.append(direction_data)

        return table_data

    def parse_direction_line(self, line):
        """Парсинг строки с данными направления с учетом структуры колонок"""
        try:
            # Убираем лишние пробелы
            cleaned_line = re.sub(r'\s+', ' ', line).strip()
            
            # Разбиваем строку по пробелам
            parts = cleaned_line.split(' ')
            
            if len(parts) < 4:
                return None
            
            # Номер направления - всегда первый элемент
            direction_number = parts[0]
            
            # Тип направления - ищем ключевые слова
            direction_type = None
            type_keywords = ['Транспортное', 'Пешеходное', 'Поворотное', 'Общ. трансп.']
            for keyword in type_keywords:
                if keyword in cleaned_line:
                    direction_type = keyword
                    break
            
            if not direction_type:
                return None
            
            # ФАЗЫ - ищем последовательность цифр с запятыми после типа направления
            phases = None
            
            # Находим позицию типа направления в строке
            type_pos = cleaned_line.find(direction_type)
            if type_pos != -1:
                # Берем подстроку после типа направления
                after_type = cleaned_line[type_pos + len(direction_type):].strip()
                
                # Ищем последовательность цифр с запятыми (формат: 1,2,3,4)
                phases_match = re.search(r'^(\d+(?:,\d+)*)', after_type)
                if phases_match:
                    phases = phases_match.group(1)
                else:
                    # Альтернативный поиск - берем первый элемент после типа
                    after_parts = after_type.split(' ')
                    if after_parts and re.match(r'^\d+(?:,\d+)*$', after_parts[0]):
                        phases = after_parts[0]
                    else:
                        # Еще один вариант - ищем любую последовательность цифр с запятыми в строке
                        all_phases_matches = re.findall(r'\b\d+(?:,\d+)+\b', cleaned_line)
                        if all_phases_matches:
                            phases = all_phases_matches[0]
            
            # Светофоры - ищем после фаз
            traffic_lights = None
            if phases:
                phases_pos = cleaned_line.find(phases)
                if phases_pos != -1:
                    # Берем подстроку после фаз
                    after_phases = cleaned_line[phases_pos + len(phases):].strip()
                    # Ищем обозначения светофоров
                    traffic_match = re.search(r'(Тр\.\s*[\d,]+|Пеш\.\s*[\d,]+|д/с\s*[\d,]+|б/л\s*[\d,]+)', after_phases)
                    if traffic_match:
                        traffic_lights = traffic_match.group(1)
                    else:
                        # Альтернативный поиск светофоров в всей строке
                        traffic_match_full = re.search(r'(Тр\.\s*[\d,]+|Пеш\.\s*[\d,]+|д/с\s*[\d,]+|б/л\s*[\d,]+)', cleaned_line)
                        if traffic_match_full:
                            traffic_lights = traffic_match_full.group(1)
            else:
                # Если фазы не найдены, все равно ищем светофоры
                traffic_match = re.search(r'(Тр\.\s*[\d,]+|Пеш\.\s*[\d,]+|д/с\s*[\d,]+|б/л\s*[\d,]+)', cleaned_line)
                if traffic_match:
                    traffic_lights = traffic_match.group(1)
            
            return {
                'number': direction_number,
                'type': direction_type,
                'phases': phases or 'не указано',
                'traffic_lights': traffic_lights or 'не указано',
            }
            
        except Exception as e:
            print(f"Ошибка парсинга строки '{line}': {e}")
        
        return None

class GenerateSwitchTable(APIView):
    """
    API для генерации таблицы коммутации контроллера
    """
    
    def post(self, request):
        """
        Генерация таблицы коммутации
        Параметры:
        - controller_type: "Поток" или "Поток-Д"
        - num_klemm: количество клемм (только для "Поток")
        - filename: имя файла (опционально)
        - passport_info: информация для паспорта (опционально)
        - table_data: данные для заполнения таблицы (опционально)
        """
        try:
            # Получаем параметры из запроса
            controller_type = request.data.get('controller_type', 'Поток')
            num_klemm = request.data.get('num_klemm', 12)
            filename = request.data.get('filename', 'Таблица ТК.xlsx')
            passport_info = request.data.get('passport_info', {})
            table_data = request.data.get('table_data', [])
            
            # Валидация параметров
            if controller_type not in ['Поток', 'Поток-Д']:
                return Response(
                    {'error': 'Неверный тип контроллера. Допустимые значения: "Поток", "Поток-Д"'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if controller_type == 'Поток':
                try:
                    num_klemm = int(num_klemm)
                    if num_klemm < 1 or num_klemm > 50:
                        return Response(
                            {'error': 'Количество клемм должно быть от 1 до 50'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                except (ValueError, TypeError):
                    return Response(
                        {'error': 'Количество клемм должно быть целым числом'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # Для Поток-Д всегда 12 клемм
                num_klemm = 12
            
            # Создаем таблицу
            wb = self.create_table(controller_type, num_klemm)
            
            # Заполняем паспортную информацию если передана
            if passport_info:
                self.fill_passport_info(wb.active, passport_info)
            
            # Заполняем данные таблицы если переданы
            if table_data:
                self.fill_table_data(wb.active, table_data, controller_type)
            
            # Сохраняем во временный файл
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                wb.save(tmp_file.name)
                tmp_file_path = tmp_file.name
            
            # Читаем файл и отправляем как ответ
            with open(tmp_file_path, 'rb') as f:
                response = HttpResponse(
                    f.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Удаляем временный файл
            os.unlink(tmp_file_path)
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Произошла ошибка при генерации таблицы: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def fill_passport_info(self, ws, passport_info):
        """Заполнение паспортной информации в заголовке таблицы"""
        number = passport_info.get('number')
        address = passport_info.get('address')
        
        if number:
            # Обновляем объект в первой строке
            ws['A1'].value = f"Светофорный объект: СО {number}"
        
        if address:
            # Обновляем адрес во второй строке
            current_value = ws['A2'].value or ""
            if "Таблица коммутации контроллера" in current_value:
                # Сохраняем тип контроллера и добавляем адрес
                controller_type = current_value.replace("Таблица коммутации контроллера", "").strip()
                ws['A2'].value = f'Таблица коммутации контроллера "{controller_type}", {address}'
    
    def fill_table_data(self, ws, table_data, controller_type):
        """Заполнение данных таблицы (№ напр., Тип напр., Разреш. фазы, № светофора)"""
        # Определяем количество строк на клемму в зависимости от типа контроллера
        rows_per_klemma = 4 if controller_type == "Поток" else 3
        
        # Создаем стиль выравнивания по центру
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for data_item in table_data:
            klemma_number = str(data_item.get('number', ''))
            if not klemma_number:
                continue
                
            # Находим строку с нужной клеммой
            klemma_row = self.find_klemma_row(ws, klemma_number, rows_per_klemma)
            if klemma_row:
                # Заполняем № направления, фазы и светофоры
                columns_data = [
                    ('E', 'number'),
                    ('G', 'phases'),
                    ('H', 'traffic_lights')
                ]
                
                for col, field in columns_data:
                    if field in data_item and data_item[field]:
                        value = data_item[field]
                        
                        # Фильтруем номер светофора - убираем буквы, точки, слеши
                        if field == 'traffic_lights':
                            value = self.filter_traffic_lights_number(value)
                        
                        cell = ws[f'{col}{klemma_row}']
                        cell.value = value
                        cell.alignment = center_alignment
                
                # Особое заполнение для типа направления
                if 'type' in data_item and data_item['type']:
                    self.fill_direction_type(ws, klemma_row, data_item['type'], controller_type)

    def fill_direction_type(self, ws, start_row, direction_type, controller_type):
        """Заполняет тип направления в зависимости от его значения"""
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Определяем количество строк в клемме
        total_rows = 4 if controller_type == "Поток" else 3
        
        direction_type = str(direction_type).strip().lower()
        
        if direction_type == 'транспортное':
            # Заполняем ТРИ строки "Транспортное"
            for i in range(3):  # Всегда 3 строки
                if i < total_rows:  # Проверяем, чтобы не выйти за границы
                    cell = ws[f'F{start_row + i}']
                    cell.value = 'Транспортное'
                    cell.alignment = center_alignment
                
        elif direction_type == 'пешеходное':
            # Заполняем ТРИ строки по шаблону
            if total_rows >= 1:  # Первая строка
                ws[f'F{start_row}'].value = 'Пешеходное'
                ws[f'F{start_row}'].alignment = center_alignment
                
            if total_rows >= 2:  # Вторая строка
                ws[f'F{start_row + 1}'].value = 'Индикация ТВП'
                ws[f'F{start_row + 1}'].alignment = center_alignment
                
            if total_rows >= 3:  # Третья строка
                ws[f'F{start_row + 2}'].value = 'Пешеходное'
                ws[f'F{start_row + 2}'].alignment = center_alignment
                
        elif direction_type == 'поворотное':
            # Заполняем только 1 и 3 строки
            if total_rows >= 1:  # Первая строка
                ws[f'F{start_row}'].value = 'Поворотное'
                ws[f'F{start_row}'].alignment = center_alignment
                
            if total_rows >= 3:  # Третья строка
                ws[f'F{start_row + 2}'].value = 'Поворотное'
                ws[f'F{start_row + 2}'].alignment = center_alignment
                
        else:
            # Для всех других типов заполняем ВСЕ строки клеммы
            original_type = str(direction_type).capitalize()
            for i in range(total_rows):  # Все строки клеммы
                cell = ws[f'F{start_row + i}']
                cell.value = original_type
                cell.alignment = center_alignment

    def filter_traffic_lights_number(self, text):
        """Фильтрует номер светофора, оставляя только цифры, запятые и дефисы"""
        if not text:
            return text
        
        # Разрешенные символы
        allowed_chars = '0123456789,- '
        
        # Фильтруем текст, оставляя только разрешенные символы
        filtered = ''.join(char for char in str(text) if char in allowed_chars)
        
        # Убираем лишние пробелы
        filtered = ' '.join(filtered.split())
        
        return filtered

    def find_klemma_row(self, ws, klemma_number, rows_per_klemma):
        """Находит начальную строку клеммы по номеру"""
        current_row = 4  # Начинаем с первой строки клемм
        
        while current_row < 1000:  # Защита от бесконечного цикла
            # Проверяем ячейку с номером клеммы (столбец A)
            cell_value = ws[f'A{current_row}'].value
            if cell_value is None:
                break
                
            if str(cell_value) == str(klemma_number):
                return current_row
            
            # Переходим к следующей клемме
            current_row += rows_per_klemma
            
            # Проверяем есть ли разделитель
            separator_check = ws[f'A{current_row}'].value
            if separator_check is not None and separator_check == '':
                current_row += 1  # Пропускаем разделитель
        
        return None

    # Все остальные методы остаются без изменений
    def create_table(self, controller_type="Поток", num_klemm=12):
        """Создает таблицу с заданным типом контроллера и количеством клемм"""
        
        wb = Workbook()
        ws = wb.active
        
        if controller_type == "Поток":
            ws.title = "Таблица ТК Поток"
            self.create_table_potok(ws, num_klemm)
        else:  # Поток-Д
            ws.title = "Таблица ТК Поток-Д"
            self.create_table_potok_d(ws)
        
        return wb

    def create_table_potok(self, ws, num_klemm):
        """Создает таблицу для контроллера Поток"""
        self.setup_worksheet(ws)
        self.create_header(ws, "Поток")
        self.create_table_headers(ws)
        self.create_klemmas_potok(ws, num_klemm)
        self.create_footer(ws, num_klemm)

    def create_table_potok_d(self, ws):
        """Создает таблицу для контроллера Поток-Д (всегда 12 клемм)"""
        self.setup_worksheet(ws)
        self.create_header(ws, "Поток-Д")
        self.create_table_headers(ws)
        self.create_klemmas_potok_d(ws)
        self.create_footer_potok_d(ws)

    def setup_worksheet(self, ws):
        """Настройка размеров столбцов и стилей"""
        column_widths = {
            'A': 6, 'B': 10, 'C': 10, 'D': 12, 'E': 12,
            'F': 20, 'G': 14, 'H': 14, 'I': 8, 'J': 8
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def create_header(self, ws, controller_type):
        """Создание заголовка таблицы"""

        ws.row_dimensions[3].height = 35
        ws.merge_cells('A1:J1')
        ws.merge_cells('A2:J2')
        
        header_data = [
            (1, "Объект: CO"),
            (2, f"Таблица коммутации контроллера {controller_type}")
        ]
        
        center = Alignment(horizontal='center', vertical='center')
        
        for row, text in header_data:
            cell = ws[f'A{row}']
            cell.value = text
            cell.alignment = center
            if row == 2:
                cell.font = Font(bold=True)

    def create_table_headers(self, ws):
        """Создание заголовков таблицы"""

        merges = ['A3:C3', 'E3:E3', 'F3:F3', 'G3:G3', 'H3:H3', 'I3:J3']
        for merge in merges:
            ws.merge_cells(merge)

        headers = [
            ('A3', '↓ Клеммы ↓'), ('D3', '№ тир-ра'), ('E3', '№ напр.'), 
            ('F3', 'Тип напр.'), ('G3', 'Разреш. фазы'), ('H3', '№ светофора'), 
            ('I3', 'Сигнал')
        ]

        bold_italic = Font(bold=True, italic=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border_bold = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        self.set_merged_cell_borders(ws, 'A3:j3', border_bold)

        for cell_ref, value in headers:
            cell = ws[cell_ref]
            cell.value = value
            cell.font = bold_italic
            cell.alignment = center
            cell.border = border_bold

    def set_merged_cell_borders(self, ws, merge_range, border):
        """Устанавливает границы для всех ячеек в объединенном диапазоне"""
        from openpyxl.worksheet.cell_range import CellRange
        cr = CellRange(merge_range)
        
        for row in range(cr.min_row, cr.max_row + 1):
            for col in range(cr.min_col, cr.max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

    def create_klemmas_potok(self, ws, num_klemm):
        """Создание клемм для Поток (4 строки на клемму)"""
        current_row = 4
        bold_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        
        for klemma_num in range(1, num_klemm + 1):
            self.create_klemma_potok(ws, current_row, klemma_num, bold_font, center, grey_fill)
            current_row += 4
            
            if (klemma_num == 2) or (klemma_num > 2 and (klemma_num - 2) % 4 == 0 and klemma_num != num_klemm):
                self.create_separator(ws, current_row, bold_font, center, blue_fill)
                current_row += 1

    def create_klemma_potok(self, ws, start_row, klemma_num, bold_font, center, grey_fill):
        """Создание одной клеммы для Поток (4 строки)"""
        merges = [
            f'A{start_row}:A{start_row+3}', f'E{start_row}:E{start_row+3}',
            f'G{start_row}:G{start_row+3}', f'H{start_row}:H{start_row+3}',
            f'I{start_row}:J{start_row}', f'I{start_row+1}:J{start_row+1}',
            f'I{start_row+2}:J{start_row+2}', f'I{start_row+3}:J{start_row+3}'
        ]
        for merge in merges:
            ws.merge_cells(merge)

        # Вычисляем номера тиристоров для этой клеммы
        # Каждая клемма имеет 4 тиристора, нумерация сквозная
        start_thyristor = (klemma_num - 1) * 4 + 1
        
        klemma_data = [
            (start_row, 'A', klemma_num),
            (start_row, 'B', 'О'),
            (start_row, 'C', ' О'),
            (start_row, 'D', start_thyristor),  # 1, 5, 9, 13 и т.д.
            (start_row, 'I', 'красный'),
            (start_row+1, 'B', 'О'),
            (start_row+1, 'C', ' О'),
            (start_row+1, 'D', start_thyristor + 1),  # 2, 6, 10, 14 и т.д.
            (start_row+1, 'I', 'желтый'),
            (start_row+2, 'B', 'О'),
            (start_row+2, 'C', ' О'),
            (start_row+2, 'D', start_thyristor + 2),  # 3, 7, 11, 15 и т.д.
            (start_row+2, 'I', 'зеленый'),
            (start_row+3, 'B', 'О'),
            (start_row+3, 'C', ' О'),
            (start_row+3, 'D', start_thyristor + 3),  # 4, 8, 12, 16 и т.д.
            (start_row+3, 'I', 'стрелка'),
        ]

        self.apply_klemma_styles(ws, klemma_data, start_row, bold_font, center, grey_fill, 4)
        
        border_middle_thin = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        self.set_merged_cell_borders(ws, f'A{start_row}:J{start_row+3}', border_middle_thin)
        self.set_merged_cell_borders(ws, f'A{start_row}:J{start_row}', 
                            Border(left=Side(style='medium'), right=Side(style='medium'), 
                                    top=Side(style='medium'), bottom=Side(style='thin')))
        self.set_merged_cell_borders(ws, f'A{start_row+3}:J{start_row+3}', 
                            Border(left=Side(style='medium'), right=Side(style='medium'), 
                                    top=Side(style='thin'), bottom=Side(style='medium')))

    def create_klemmas_potok_d(self, ws):
        """Создание клемм для Поток-Д (3 строки на клемму, разделитель после каждой)"""
        current_row = 4
        bold_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        
        for klemma_num in range(1, 13):  # Всегда 12 клемм для Поток-Д
            self.create_klemma_potok_d(ws, current_row, klemma_num, bold_font, center, grey_fill)
            current_row += 3
            
            # Разделитель после КАЖДОЙ клеммы, включая последнюю
            self.create_separator(ws, current_row, bold_font, center, blue_fill)
            current_row += 1

    def create_klemma_potok_d(self, ws, start_row, klemma_num, bold_font, center, grey_fill):
        """Создание одной клеммы для Поток-Д (3 строки)"""
        merges = [
            f'A{start_row}:A{start_row+2}', f'E{start_row}:E{start_row+2}',
            f'G{start_row}:G{start_row+2}', f'H{start_row}:H{start_row+2}',
            f'I{start_row}:J{start_row}', f'I{start_row+1}:J{start_row+1}',
            f'I{start_row+2}:J{start_row+2}'
        ]
        for merge in merges:
            ws.merge_cells(merge)

        # Вычисляем номера тиристоров для этой клеммы
        # Каждая клемма имеет 3 тиристора, нумерация сквозная
        start_thyristor = (klemma_num - 1) * 3 + 1

        klemma_data = [
            (start_row, 'A', klemma_num),
            (start_row, 'B', 'О'),
            (start_row, 'C', ' О'),
            (start_row, 'D', start_thyristor),  # 1, 4, 7, 10 и т.д.
            (start_row, 'I', 'красный'),
            (start_row+1, 'B', 'О'),
            (start_row+1, 'C', ' О'),
            (start_row+1, 'D', start_thyristor + 1),  # 2, 5, 8, 11 и т.д.
            (start_row+1, 'I', 'желтый'),
            (start_row+2, 'B', 'О'),
            (start_row+2, 'C', ' О'),
            (start_row+2, 'D', start_thyristor + 2),  # 3, 6, 9, 12 и т.д.
            (start_row+2, 'I', 'зеленый'),
        ]

        self.apply_klemma_styles(ws, klemma_data, start_row, bold_font, center, grey_fill, 3)
        
        border_middle_thin = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        self.set_merged_cell_borders(ws, f'A{start_row}:J{start_row+2}', border_middle_thin)
        self.set_merged_cell_borders(ws, f'A{start_row}:J{start_row}', 
                            Border(left=Side(style='medium'), right=Side(style='medium'), 
                                    top=Side(style='medium'), bottom=Side(style='thin')))
        self.set_merged_cell_borders(ws, f'A{start_row+2}:J{start_row+2}', 
                            Border(left=Side(style='medium'), right=Side(style='medium'), 
                                    top=Side(style='thin'), bottom=Side(style='medium')))

    def apply_klemma_styles(self, ws, klemma_data, start_row, bold_font, center, grey_fill, rows_per_klemma):
        """Применение стилей к ячейкам клеммы"""
        border_top_bold = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='medium'), bottom=Side(style='thin')
        )
        
        border_middle_thin = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        border_bottom_bold = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='thin'), bottom=Side(style='medium')
        )
        
        for row, col, value in klemma_data:
            cell = ws[f'{col}{row}']
            cell.value = value
            cell.alignment = center
            
            if row == start_row:
                if col in ['B', 'C', 'D']:
                    cell.border = border_top_bold
            elif row == start_row + rows_per_klemma - 1:
                if col in ['B', 'C', 'D']:
                    cell.border = border_bottom_bold
            else:
                if col in ['B', 'C', 'D']:
                    cell.border = border_middle_thin
            
            if col in ['A', 'B', 'C']:
                cell.font = bold_font
            
            if col in ['B', 'C']:
                cell.fill = grey_fill

    def create_separator(self, ws, row, bold_font, center, blue_fill):
        """Создание разделителя между клеммами"""
        ws.merge_cells(f'D{row}:H{row}')
        ws.merge_cells(f'I{row}:J{row}')

        separator_data = [
            (f'A{row}', '', blue_fill),
            (f'B{row}', 'О', blue_fill),
            (f'C{row}', ' О', blue_fill),
            (f'D{row}', '', None),  
            (f'I{row}', 'Общий', None),
        ]

        border_bold = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='medium'), bottom=Side(style='medium')
        )

        for cell_ref, value, fill in separator_data:
            cell = ws[cell_ref]
            cell.value = value
            cell.alignment = center
            cell.border = border_bold
            if fill:
                cell.fill = fill

        self.set_merged_cell_borders(ws, f'A{row}:J{row}', border_bold)

    def create_footer(self, ws, num_klemm):
        """Создание завершающей части таблицы для Поток"""
        for row in range(7, 1000):
            if ws[f'B{row}'].value is None:
                footer_row = row
                break

        ws.merge_cells(f'A{footer_row}:A{footer_row+2}')
        ws.merge_cells(f'G{footer_row}:J{footer_row}')
        ws.merge_cells(f'G{footer_row+1}:J{footer_row+1}')
        ws.merge_cells(f'C{footer_row+2}:J{footer_row+2}')

        center = Alignment(horizontal='center', vertical='center')

        footer_data = [
            (footer_row, [
                ('B', 'О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
                ('C', ' О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
                ('E', '─', None, center),
                ('F', 'фаза', None, center),
                ('G', 'Вых.220в для подключения УЗСП, ТООВ и т.д.', None, Alignment(vertical='center'))
            ]),
            (footer_row+1, [
                ('B', 'О', PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), center),
                ('C', 'О', PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), center),
                ('E', '─', None, center),
                ('F', 'земля', None, center),
                ('G', 'Заземление', None, Alignment(vertical='center'))
            ])
        ]
        
        self.apply_footer_styles(ws, footer_data, footer_row)
        
        border_bold = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        self.set_merged_cell_borders(ws, f'A{footer_row}:J{footer_row+2}', border_bold)

    def create_footer_potok_d(self, ws):
        """Создание завершающей части таблицы для Поток-Д"""
        for row in range(7, 1000):
            if ws[f'B{row}'].value is None:
                footer_row = row
                break

        ws.merge_cells(f'A{footer_row}:A{footer_row+2}')
        ws.merge_cells(f'G{footer_row}:J{footer_row}')
        ws.merge_cells(f'G{footer_row+1}:J{footer_row+1}')
        ws.merge_cells(f'C{footer_row+2}:J{footer_row+2}')

        center = Alignment(horizontal='center', vertical='center')

        footer_data = [
            (footer_row, [
                ('B', 'О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
                ('C', ' О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
                ('E', '─', None, center),
                ('F', 'фаза', None, center),
                ('G', 'Вых.220в для подключения УЗСП, ТООВ и т.д.', None, Alignment(vertical='center'))
            ]),
            (footer_row+1, [
                ('B', 'О', PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), center),
                ('C', 'О', PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), center),
                ('E', '─', None, center),
                ('F', 'земля', None, center),
                ('G', 'Заземление', None, Alignment(vertical='center'))
            ])
        ]
        
        self.apply_footer_styles(ws, footer_data, footer_row)
        
        border_bold = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        self.set_merged_cell_borders(ws, f'A{footer_row}:J{footer_row+2}', border_bold)

    def apply_footer_styles(self, ws, footer_data, start_row):
        """Применение стилей к футеру"""
        bold_font = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        
        border_bold = Border(
            left=Side(style='medium'), right=Side(style='medium'), 
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        for row_offset, cells in footer_data:
            for col, value, fill, alignment in cells:
                cell = ws[f'{col}{row_offset}']
                cell.value = value
                cell.font = bold_font
                cell.alignment = alignment if alignment else center
                cell.border = border_bold
                
                if fill:
                    cell.fill = fill
