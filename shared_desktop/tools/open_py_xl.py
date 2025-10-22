from django.shortcuts import redirect, render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from openpyxl import load_workbook, Workbook
from io import BytesIO
from tools.models import PhaseParameterSet
from tools.serializers import PhaseParameterSetSerializer
from collections import defaultdict
from datetime import datetime, timedelta


class OpenpyxlAPIView(APIView):
    def get(self, request, name=None):
        if name:
            try:
                selected_set = PhaseParameterSet.objects.get(name=name)
                serializer = PhaseParameterSetSerializer(selected_set)
                return Response(serializer.data)
            except PhaseParameterSet.DoesNotExist:
                return Response({"error": "Набор параметров не найден."}, status=status.HTTP_404_NOT_FOUND)
        else:
            parameter_sets = PhaseParameterSet.objects.all()
            serializer = PhaseParameterSetSerializer(parameter_sets, many=True)
            return Response(serializer.data)

    def post(self, request):
        print("Входящие данные (request.data):")
        for key, value in request.data.items():
            if hasattr(value, 'read'):  # Проверяем, является ли значение файлом
                print(f"{key}: <InMemoryUploadedFile>")
            else:
                print(f"{key}: {value}")
        print("Входящие файлы (request.FILES):", request.FILES)
        if 'upload_button' in request.data:
            selected_set_name = request.data.get("parameter_set")
            try:
                selected_set = PhaseParameterSet.objects.get(name=selected_set_name)
                serializer = PhaseParameterSetSerializer(selected_set)
                return Response(serializer.data)
            except PhaseParameterSet.DoesNotExist:
                return Response({"error": "Выбранные настройки не найдены."}, status=status.HTTP_404_NOT_FOUND)

        elif 'save_button' in request.data:
            name = request.data.get("name")
            base_name = name
            counter = 1
            while PhaseParameterSet.objects.filter(name=name).exists():
                name = f"{base_name}_{counter}"
                counter += 1

            primary_group = request.data.get("primary_group")
            groups = {f"group{i}": request.data.get(f"group{i}") for i in range(1, 21)}
            phases = {f"phases{i}": request.data.get(f"phases{i}") for i in range(1, 21)}

            data = {"primary_group": primary_group, **groups, **phases}
            PhaseParameterSet.objects.create(name=name, data=data)
            return Response({"status": "saved", "name": name}, status=status.HTTP_201_CREATED)

        elif 'process_button' in request.data:
            if 'file' in request.FILES:
                try:
                    excel_file = request.FILES['file']
                    workbook = load_workbook(excel_file)
                    sheet = workbook.active
                    sheet.title = "Данные"
                    primary_group = request.data.get("primary_group")
                    interval_minutes = int(request.data.get("interval", 60))
                    group_mapping = {f"group{i}": request.data.get(f"group{i}") for i in range(1, 21)}

                    rows = list(sheet.iter_rows(min_row=2, values_only=True))
                    sorted_rows = sorted(rows, key=lambda row: row[0] if row[0] is not None else '')
                    new_rows = [split_record_by_minute(row) for row in sorted_rows]
                    sorted_rows = [item for sublist in new_rows for item in sublist]

                    phases = {}
                    for i in range(1, 21):
                        group = request.data.get(f"group{i}")
                        phase = request.data.get(f"phases{i}")
                        if phase:
                            phase_list = [p.strip() for p in phase.split(',')]
                            for p in phase_list:
                                if p not in phases:
                                    phases[p] = []
                                if group:
                                    phases[p].append(group)

                    phases['-1'] = ['-1']

                    sheet.cell(row=1, column=4, value="Направления")
                    sheet.cell(row=1, column=5, value="Общая длительность")
                    sheet.cell(row=1, column=6, value="Счетчик циклов")
                    sheet.cell(row=1, column=7, value="Циклы за час")
                    sheet.column_dimensions['D'].width = 20
                    sheet.column_dimensions['E'].width = 20
                    sheet.column_dimensions['F'].width = 18
                    sheet.column_dimensions['G'].width = 15

                    sheet = calculate_total_duration(sheet, sorted_rows, phases)
                    sheet = calculate_cycle_count(sheet, sorted_rows, phases, primary_group, group_mapping)
                    sheet = calculate_cycles_per_hour(sheet, sorted_rows, interval_minutes)
                    sheet = calculate_total_duration_per_group(sheet, sorted_rows, group_mapping, interval_minutes)

                    workbook = create_filtered_sheet_and_transfer_data(sheet, interval_minutes)
                    workbook = create_filtered_sheet_and_transfer_data_2(sheet)

                    for i in range(min(3, len(workbook.worksheets))):
                        workbook.worksheets[i].freeze_panes = "A2"

                    virtual_workbook = BytesIO()
                    workbook.save(virtual_workbook)
                    virtual_workbook.seek(0)

                    response = HttpResponse(virtual_workbook, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="openpyxl.xlsx"'
                    return response
                except Exception as e:
                    print("Ошибка при обработке файла:", str(e))
                    return Response({"error": "Ошибка при обработке файла."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            else:
                return Response({"error": "Файл не найден."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"error": "Неизвестная операция."}, status=status.HTTP_400_BAD_REQUEST)


def openpyxl(request):
    parameter_sets = PhaseParameterSet.objects.all()
    selected_set = None
    data = {}

    if request.method == "POST":
        if "upload_button" in request.POST:
            selected_set_name = request.POST.get("parameter_set")

            if not selected_set_name:
                print("Пожалуйста, выберите сохраненные настройки.")
            else:
                try:
                    selected_set = PhaseParameterSet.objects.get(name=selected_set_name)
                    data = selected_set.data
                except PhaseParameterSet.DoesNotExist:
                    print("Выбранные настройки не найдены.")

            context = {
                "parameter_sets": parameter_sets,
                "selected_set": selected_set,
                "data": data,
                "selected_set_name": selected_set_name,
            }
            return render(request, "tools/openpyxl.html", context)

        elif "save_button" in request.POST:
            name = request.POST.get("name")
            base_name = name
            counter = 1

            while PhaseParameterSet.objects.filter(name=name).exists():
                name = f"{base_name}_{counter}"
                counter += 1

            primary_group = request.POST.get("primary_group")
            groups = {f"group{i}": request.POST.get(f"group{i}") for i in range(1, 21)}
            phases = {f"phases{i}": request.POST.get(f"phases{i}") for i in range(1, 21)}

            data = {"primary_group": primary_group, **groups, **phases}

            PhaseParameterSet.objects.create(name=name, data=data)
            return redirect("/tools/openpyxl/")

        elif "process_button" in request.POST:
            if 'file' in request.FILES:
                excel_file = request.FILES['file']
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                sheet.title = "Данные"
                primary_group = request.POST.get("primary_group")
                interval_minutes = int(request.POST.get("interval", 60))
                # Создаем маппинг для групп
                group_mapping = {f"group{i}": request.POST.get(f"group{i}") for i in range(1, 21)}

                rows = list(sheet.iter_rows(min_row=2, values_only=True))
                sorted_rows = sorted(rows, key=lambda row: row[0] if row[0] is not None else '')

                # Разбиваем записи на промежуточные
                new_rows = []
                for row in sorted_rows:
                    new_rows.extend(split_record_by_minute(row))

                # Обновляем sorted_rows
                sorted_rows = new_rows

                phases = {}
                for i in range(1, 21):
                    group = request.POST.get(f"group{i}")
                    phase = request.POST.get(f"phases{i}")
                    if phase:
                        phase_list = [p.strip() for p in phase.split(',')]
                        for p in phase_list:
                            if p not in phases:
                                phases[p] = []
                            if group:
                                phases[p].append(group)

                # Добавляем направление 'нет связи'
                phases['-1'] = ['-1']

                sheet.cell(row=1, column=4, value="Направления")
                sheet.cell(row=1, column=5, value="Общая длительность")
                sheet.cell(row=1, column=6, value="Счетчик циклов")
                sheet.cell(row=1, column=7, value="Циклы за час")
                sheet.column_dimensions['D'].width = 20
                sheet.column_dimensions['E'].width = 20
                sheet.column_dimensions['F'].width = 18
                sheet.column_dimensions['G'].width = 15

                # Вызываем методы для подсчета общей длительности, счетчика циклов и циклов за час
                sheet = calculate_total_duration(sheet, sorted_rows, phases)
                sheet = calculate_cycle_count(sheet, sorted_rows, phases, primary_group, group_mapping)
                sheet = calculate_cycles_per_hour(sheet, sorted_rows, interval_minutes)
                sheet = calculate_total_duration_per_group(sheet, sorted_rows, group_mapping, interval_minutes)

                # Вызываем метод для создания нового листа и переноса данных
                workbook = create_filtered_sheet_and_transfer_data(sheet, interval_minutes)
                workbook = create_filtered_sheet_and_transfer_data_2(sheet)

                # Закрепляем первую строку на всех листах
                for i in range(min(3, len(workbook.worksheets))):
                    workbook.worksheets[i].freeze_panes = "A2"

                virtual_workbook = BytesIO()
                workbook.save(virtual_workbook)
                virtual_workbook.seek(0)

                response = HttpResponse(virtual_workbook, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="openpyxl.xlsx"'
                return response

    context = {
        "parameter_sets": parameter_sets,
        "selected_set": selected_set,
        "data": data,
    }
    return render(request, "tools/openpyxl.html", context)

def split_record_by_minute(record):
    start_time, phase, duration = record
    start_time = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S") if isinstance(start_time, str) else start_time
    records = []
    
    while duration > 0:
        next_minute = (start_time.replace(second=0, microsecond=0)) + timedelta(minutes=1)
        time_diff = (next_minute - start_time).total_seconds()
        
        if duration <= time_diff:
            # Если оставшаяся длительность меньше или равна разнице до следующей минуты
            if duration > 1:
                # Разделяем на две записи: основную часть и последнюю секунду
                records.append((start_time, phase, round(duration - 1)))
                records.append((start_time + timedelta(seconds=duration - 1), phase, 1))
            else:
                # Если длительность 1 секунда, просто добавляем одну запись
                records.append((start_time, phase, 1))
            break
        else:
            # Разделяем текущую минуту на две записи
            # Первая запись: от start_time до next_minute - 1 секунда
            records.append((start_time, phase, round(time_diff - 1)))
            # Вторая запись: последняя секунда минуты
            records.append((next_minute - timedelta(seconds=1), phase, 1))
            
            # Обновляем start_time и уменьшаем duration
            duration -= time_diff
            start_time = next_minute
    
    return records

def calculate_total_duration(sheet, sorted_rows, phases):
    current_direction = None
    current_sum_duration = 0

    for idx, row in enumerate(sorted_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=idx, column=col_idx, value=value)

        phase_value = sheet.cell(row=idx, column=2).value
        direction_cell = sheet.cell(row=idx, column=4)
        duration_cell = sheet.cell(row=idx, column=3)

        # Устанавливаем значение в столбец D
        directions = phases.get(str(phase_value), [])
        direction_cell.value = ", ".join(directions) if directions else ""

        # Получаем значение длительности из столбца C
        try:
            duration_value = float(duration_cell.value) if duration_cell.value is not None else 0
        except ValueError:
            duration_value = 0

        # Суммируем длительности для одинаковых направлений
        if direction_cell.value == current_direction:
            current_sum_duration += duration_value
        else:
            if current_direction:
                # Записываем сумму в последней строке группы
                sheet.cell(row=idx-1, column=5, value=current_sum_duration)

            # Обновляем текущие значения
            current_direction = direction_cell.value
            current_sum_duration = duration_value
            first_row_of_group = idx

    # Записываем сумму длительности в последней строке группы
    if current_direction:
        sheet.cell(row=idx, column=5, value=current_sum_duration)

    return sheet

def calculate_cycle_count(sheet, sorted_rows, phases, primary_group, group_mapping):
    cycle_counter = 0
    previous_phase = None
    primary_group_active = False

    for idx, row in enumerate(sorted_rows, start=2):
        phase_value = sheet.cell(row=idx, column=2).value
        direction_cell = sheet.cell(row=idx, column=4)
        duration_cell = sheet.cell(row=idx, column=3)

        if primary_group:
            primary_group_name = group_mapping.get(primary_group)
            direction_value = direction_cell.value

            # Проверяем, если текущая фаза принадлежит primary_group
            current_primary_group_active = primary_group_name in direction_value.split(", ")

            # Увеличиваем счетчик циклов, если была смена фазы и текущая фаза не в primary_group
            if previous_phase and previous_phase != phase_value:
                if not primary_group_active and current_primary_group_active:
                    cycle_counter += 1
                elif '-1' in direction_value:
                    cycle_counter += 1

            # Обновляем состояние primary_group
            primary_group_active = current_primary_group_active
        else:
            # Увеличиваем счетчик циклов, если направление 'нет связи'
            if '-1' in direction_cell.value:
                cycle_counter += 1

        previous_phase = phase_value
        sheet.cell(row=idx, column=6, value=cycle_counter)

    return sheet

def calculate_cycles_per_hour(sheet, sorted_rows, interval_minutes=60):
    last_seen_time = None  # Последнее обработанное время
    starting_cycle = None  # Начальное значение счетчика циклов для текущего интервала
    ending_cycle = None    # Конечное значение счетчика циклов для текущего интервала

    for idx, row in enumerate(sorted_rows, start=2):
        time_cell = sheet.cell(row=idx, column=1).value  # Время из столбца 1
        current_cycle = sheet.cell(row=idx, column=6).value  # Счетчик циклов из столбца 6

        if time_cell:
            # Округляем время до ближайшего интервала
            current_time = time_cell.replace(second=0, microsecond=0)
            interval = timedelta(minutes=interval_minutes)
            rounded_time = (current_time.replace(minute=0) + timedelta(minutes=(current_time.minute // interval_minutes)) * interval_minutes)

            # Если время изменилось на новый интервал
            if rounded_time != last_seen_time:
                if last_seen_time is not None:
                    # Подсчет циклов за предыдущий интервал
                    if starting_cycle is not None and ending_cycle is not None:
                        cycles_per_interval = ending_cycle - starting_cycle
                        sheet.cell(row=idx - 1, column=7, value=cycles_per_interval)

                # Обновляем для нового интервала
                starting_cycle = current_cycle
                last_seen_time = rounded_time

            # Обновляем конечное значение счетчика циклов
            ending_cycle = current_cycle

    # Обработка последнего интервала
    if last_seen_time is not None and starting_cycle is not None and ending_cycle is not None:
        cycles_per_interval = ending_cycle - starting_cycle
        sheet.cell(row=idx, column=7, value=cycles_per_interval)

    return sheet

def calculate_total_duration_per_group(sheet, sorted_rows, group_mapping, interval_minutes):
    interval_duration = defaultdict(lambda: defaultdict(float))  # Хранение данных по интервалам
    last_seen_interval = None  # Последний обработанный интервал
    group_columns = {}  # Столбцы для групп

    # Создаем маппинг столбцов для групп только если есть группы
    header_row = 1
    col_idx = 8  # Начало столбцов для групп
    for group in group_mapping.values():
        if group.strip():  # Убедимся, что группа не пустая
            group_columns[group] = col_idx
            sheet.cell(row=header_row, column=col_idx, value=group)
            sheet.column_dimensions[sheet.cell(row=header_row, column=col_idx).column_letter].width = 8
            col_idx += 1

    for idx, row in enumerate(sorted_rows, start=2):
        time_cell = sheet.cell(row=idx, column=1).value  # Время из столбца 1
        duration_value = sheet.cell(row=idx, column=3).value  # Длительность из столбца 3
        direction_cell = sheet.cell(row=idx, column=4).value  # Направления из столбца 4

        if time_cell:
            # Округляем время до ближайшего интервала
            interval = timedelta(minutes=interval_minutes)
            current_time = time_cell.replace(second=0, microsecond=0)
            rounded_time = (current_time.replace(minute=0) + 
                            timedelta(minutes=(current_time.minute // interval_minutes)) * interval_minutes)

            # Если интервал изменился
            if rounded_time != last_seen_interval:
                # Записываем суммы длительностей за предыдущий интервал
                if last_seen_interval is not None:
                    for group, col_idx in group_columns.items():
                        sheet.cell(row=idx - 1, column=col_idx, value=interval_duration[last_seen_interval][group])

                # Обновляем для нового интервала
                last_seen_interval = rounded_time
                interval_duration[last_seen_interval] = defaultdict(float)

            # Суммируем длительности для групп
            if direction_cell:
                groups = [group.strip() for group in direction_cell.split(',')]
                for group in groups:
                    if group in group_columns:
                        interval_duration[rounded_time][group] += duration_value

    # Записываем длительности за последний интервал
    if last_seen_interval:
        for group, col_idx in group_columns.items():
            sheet.cell(row=idx, column=col_idx, value=interval_duration[last_seen_interval][group])

    # Переименовываем столбцы после всех вычислений
    for group, col_idx in group_columns.items():
        if isinstance(group, str) and group.isdigit():
            new_group_name = f"{group} н"
            sheet.cell(row=header_row, column=col_idx, value=new_group_name)
    
    # Изменяем ширину столбцов после переименования
    for group, col_idx in group_columns.items():
        new_group_name = sheet.cell(row=header_row, column=col_idx).value

        # Проверяем длину нового имени и изменяем ширину столбца
        if len(new_group_name) <= 4:
            sheet.column_dimensions[sheet.cell(row=header_row, column=col_idx).column_letter].width = 6
        else:
            sheet.column_dimensions[sheet.cell(row=header_row, column=col_idx).column_letter].width = 16

    return sheet

def create_filtered_sheet_and_transfer_data(sheet, interval_minutes):
    """Создание нового листа и перенос данных с округлением времени к следующему часу."""

    # Создаем новый лист в рабочей книге с заданным названием
    wb = sheet.parent  # Получаем рабочую книгу из текущего листа
    new_sheet_name = "Время работы направлений за час"

    # Проверяем, существует ли уже лист с таким названием и удаляем его, если существует
    if new_sheet_name in wb.sheetnames:
        std = wb[new_sheet_name]
        wb.remove(std)

    new_sheet = wb.create_sheet(title=new_sheet_name)

    # Определяем максимальное количество строк и столбцов в исходном листе
    max_row = sheet.max_row
    max_column = sheet.max_column

    # Копируем заголовки в новый лист
    for col in range(1, max_column + 1):
        # Копируем заголовки
        new_sheet.cell(row=1, column=col, value=sheet.cell(row=1, column=col).value)

    # Переименовываем столбец A в "Период"
    new_sheet.cell(row=1, column=1, value="Период")

    # Устанавливаем ширину столбцов
    new_sheet.column_dimensions['A'].width = 26
    new_sheet.column_dimensions['B'].width = 13

    # Начинаем со второй строки (первая строка - заголовки)
    new_row_index = 2

    for row_index in range(2, max_row + 1):
        if sheet.cell(row=row_index, column=7).value:  # Проверяем, есть ли данные в столбце G
            end_time = sheet.cell(row=row_index, column=1).value

            if isinstance(end_time, datetime):
                rounded_minute = (end_time.minute // interval_minutes) * interval_minutes
                start_time = end_time.replace(second=0, microsecond=0, minute=rounded_minute)
                end_time = start_time + timedelta(minutes=int(interval_minutes))

                # Форматируем строку периода
                period_str = f"{start_time.strftime('%d-%m-%Y с %H:%M')} до {end_time.strftime('%H:%M')}"

                # Записываем форматированный период в первый столбец нового листа
                new_sheet.cell(row=new_row_index, column=1, value=period_str)

            # Переносим остальные столбцы
            for col_index in range(2, max_column + 1):
                new_sheet.cell(row=new_row_index, column=col_index, value=sheet.cell(row=row_index, column=col_index).value)
            
            new_row_index += 1

    # Удаляем столбцы B, C, D, E, F
    new_sheet.delete_cols(2, 5)  # Удаляем 5 столбцов, начиная с 2-го (B)

    # Устанавливаем ширину столбцов начиная с C, в зависимости от длины названия
    for col in range(3, max_column + 1):
        header_value = new_sheet.cell(row=1, column=col).value  # Получаем название столбца

        # Проверяем длину названия столбца
        col_letter = chr(64 + col)
        if header_value and len(header_value) <= 4:
            new_sheet.column_dimensions[col_letter].width = 6
        else:
            new_sheet.column_dimensions[col_letter].width = 16

    return wb

def create_filtered_sheet_and_transfer_data_2(sheet):
    """Создание нового листа с расчетом среднего времени за цикл."""

    # Получаем рабочую книгу из текущего листа
    wb = sheet.parent  # sheet — это лист из первого метода
    source_sheet = wb.worksheets[1]  # Второй лист (индексация начинается с 0)
    new_sheet_name = "Среднее время за цикл"

    # Проверяем, существует ли уже лист с таким названием и удаляем его, если существует
    if new_sheet_name in wb.sheetnames:
        std = wb[new_sheet_name]
        wb.remove(std)

    new_sheet = wb.create_sheet(title=new_sheet_name)

    # Определяем максимальное количество строк и столбцов в исходном листе
    max_row = source_sheet.max_row
    max_column = source_sheet.max_column

    # Копируем заголовки в новый лист
    for col in range(1, max_column + 1):
        new_sheet.cell(row=1, column=col, value=source_sheet.cell(row=1, column=col).value)

    # Переименовываем столбец A в "Период"
    new_sheet.cell(row=1, column=1, value="Период")

    # Устанавливаем ширину столбцов
    new_sheet.column_dimensions['A'].width = 26

    # Устанавливаем ширину столбцов начиная с B равной 6
    for col in range(2, max_column + 1):
        col_letter = chr(64 + col)
        new_sheet.column_dimensions[col_letter].width = 6

    # Копируем столбец "Период" из исходного листа
    for row_index in range(2, max_row + 1):
        new_sheet.cell(row=row_index, column=1, value=source_sheet.cell(row=row_index, column=1).value)

    # Переносим значения из исходного листа и вычисляем средние значения
    for col_index in range(2, max_column + 1):
        group_name = source_sheet.cell(row=1, column=col_index).value
        new_sheet.cell(row=1, column=col_index, value=group_name)

        for row_index in range(2, max_row + 1):
            value = source_sheet.cell(row=row_index, column=col_index).value
            num_cycles = source_sheet.cell(row=row_index, column=2).value  # Значение из столбца B

            # Проверяем, что значение из столбца B не ноль и значение в ячейке существует
            if num_cycles and num_cycles != 0:
                try:
                    numeric_value = float(value)
                    average_time = int(numeric_value / num_cycles)
                    new_sheet.cell(row=row_index, column=col_index, value=average_time)
                except (ValueError, TypeError):
                    new_sheet.cell(row=row_index, column=col_index, value=value)
            else:
                # Если нет данных для деления, оставляем исходное значение или пусто
                new_sheet.cell(row=row_index, column=col_index, value=value)
    
    # Удаляем столбец B
    new_sheet.delete_cols(2, 1)
    # Устанавливаем ширину столбцов начиная с B, в зависимости от длины названия
    for col in range(2, max_column + 1):
        header_value = new_sheet.cell(row=1, column=col).value  # Получаем название столбца

        # Проверяем длину названия столбца
        col_letter = chr(64 + col)
        if header_value and len(header_value) <= 4:
            new_sheet.column_dimensions[col_letter].width = 6
        else:
            new_sheet.column_dimensions[col_letter].width = 16

    return wb
