import requests
import openpyxl
from tkinter import filedialog
from tkinter import Tk
from ping3 import ping
import time
import re

"""
=== ОПИСАНИЕ ТРЕБОВАНИЙ К EXCEL-ФАЙЛУ ===

Файл должен иметь следующую структуру:

| A (Любые данные) | B (Любые данные) | C (IP адрес)     | D (Тип ДК) | E (Версия) | F (Статус) |
|------------------|------------------|------------------|------------|------------|------------|
| ...              | ...              | 192.168.1.1      | Swarco     |            |            |
| ...              | ...              | 10.10.10.2       | Поток (P)  |            |            |
| ...              | ...              | 172.16.0.3       | Поток (S)  |            |            |

ОБЯЗАТЕЛЬНЫЕ ТРЕБОВАНИЯ:
1. IP-адреса должны быть в столбце C (3-й столбец)
2. Типы протокола (Тип ДК) должны быть в столбце D (4-й столбец)
3. Данные должны начинаться со 2-й строки (1-я строка - заголовки)

ЧТО СКРИПТ СДЕЛАЕТ:
- Заполнит столбец E версией прошивки или сообщением об ошибке
- Заполнит столбец F статусом обработки
- Скрипт пропустит уже обработанные строки, обработает только ошибочные или пустые

ПРИМЕР РЕЗУЛЬТАТА:
| C (IP)       | D (Тип ДК) | E (Версия) | F (Статус)            |
|--------------|------------|------------|-----------------------|
| 192.168.1.1  | Swarco     | v2.1.5     | успешно               |
| 10.10.10.2   | Поток (P)  | нет связи  | нет ping              |
| 172.16.0.3   | Поток (S)  | v1.0.0     | было обработано ранее |
"""

# Функция для проверки валидности IP-адреса
def is_valid_ip(ip):
    if not ip or not isinstance(ip, str):
        return False
    
    # Проверяем, начинается ли IP на запрещенные паттерны
    if ip.startswith('0.') or ip.startswith('192.168.'):
        return False
    
    # Регулярное выражение для проверки формата IP-адреса
    ip_pattern = re.compile(r'^(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})$')
    
    if not ip_pattern.match(ip):
        return False
    
    # Проверяем, что каждый октет в диапазоне 0-255
    octets = ip.split('.')
    for octet in octets:
        if not octet.isdigit() or not (0 <= int(octet) <= 255):
            return False
    
    return True

# Функция для отправки запроса к API и получения версии прошивки
def get_firmware(ip, protocol):
    url = "http://192.168.45.116/get_firmware_api/"
    start_time = time.time()
    max_wait_time = 12

    data = {
        "ip": ip,
        "protocol": protocol
    }

    print(f"Отправляем запрос для IP: {ip}, Протокол: '{protocol}'")  # ДЛЯ ОТЛАДКИ

    try:
        attempt = 0
        while True:
            attempt += 1
            if time.time() - start_time > max_wait_time:
                print(f"Превышено время ожидания для {ip}")
                return "Таймаут ожидания"
            
            response = requests.post(url, json=data, timeout=20)
            print(f"Попытка {attempt}: Status Code: {response.status_code}")  # ДЛЯ ОТЛАДКИ
            
            if response.status_code == 200:
                response_data = response.json()
                print(f"Ответ сервера: {response_data}")  # ДЛЯ ОТЛАДКИ
                
                version = response_data.get("version") or response_data.get("errorMessage")

                if version == "loading":
                    print(f"Версия для {ip} все еще в процессе загрузки, ждем...")
                    time.sleep(1)
                    continue
                else:
                    print(f"Получена версия: {version}")  # ДЛЯ ОТЛАДКИ
                    return version
            else:
                print(f"Ошибка HTTP: {response.status_code}")
                return f"Ошибка: {response.status_code}"
                
    except requests.Timeout:
        print(f"Тайм-аут для {ip}")
        return "Тайм-аут запроса"
    except Exception as e:
        print(f"Исключение: {str(e)}")
        return f"Ошибка при запросе: {str(e)}"

# Функция для проверки пинга
def check_ping(ip):
    try:
        # Проверяем пинг с таймаутом 2 секунды
        response = ping(ip, timeout=2)
        return response is not None
    except Exception as e:
        print(f"Ошибка при пинге {ip}: {str(e)}")
        return False

# Функция для проверки, нужно ли перепроверять строку
def should_recheck(existing_version):
    if not existing_version:
        return True
    if str(existing_version).startswith("Поток") or str(existing_version).startswith("ITC"):
        return False
    return True

# Функция для обработки Excel файла
def process_excel(file_path):
    # Загружаем Excel файл
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # Выбираем активный лист

    # Добавим заголовки для новых столбцов (если их еще нет)
    if sheet.cell(row=1, column=5).value is None:
        sheet.cell(row=1, column=5, value="Версия")
    if sheet.cell(row=1, column=6).value is None:
        sheet.cell(row=1, column=6, value="Статус")

    # Статистика
    total_processed = 0
    success_count = 0
    no_connection_count = 0
    invalid_ip_count = 0
    previously_processed_correct = 0
    previously_processed_incorrect = 0
    error_count = 0
    rechecked_count = 0

    # Проходим по строкам начиная с 2-й (пропускаем заголовки)
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=6):  # max_col=6 чтобы проверять столбец F
        ip = row[0].value  # Столбец C (IP)
        protocol = row[1].value  # Столбец D (Тип ДК)
        existing_version = row[2].value  # Столбец E (Версия) - это 3-й элемент в кортеже (индекс 2)

        total_processed += 1

        # Проверяем, нужно ли перепроверять строку
        if existing_version is not None and not should_recheck(existing_version):
            print(f"Строка {row[0].row} обработано ранее ({existing_version}), пропускаем.")
            sheet.cell(row=row[0].row, column=6, value="обработано ранее")
            previously_processed_correct += 1
            continue

        # Если есть существующая версия, но она не правильная - перепроверяем
        if existing_version is not None:
            print(f"Строка {row[0].row} имеет версию '{existing_version}', перепроверяем...")
            rechecked_count += 1
            # Очищаем старые значения для перепроверки
            sheet.cell(row=row[0].row, column=5, value=None)  # Версия (столбец E)
            sheet.cell(row=row[0].row, column=6, value=None)  # Статус (столбец F)

        if ip and protocol:
            # Проверяем валидность IP-адреса
            if not is_valid_ip(str(ip).strip()):
                print(f"Невалидный IP-адрес: {ip}")
                sheet.cell(row=row[0].row, column=5, value="проверьте IP адрес")
                sheet.cell(row=row[0].row, column=6, value="ошибка")
                invalid_ip_count += 1
                continue
                
            print(f"Обрабатываем: IP = {ip}, Тип ДК = {protocol}")

            # Проверяем пинг перед отправкой запроса
            ping_result = check_ping(ip)
            if not ping_result:
                # Если пинг не прошел, записываем "нет связи"
                sheet.cell(row=row[0].row, column=5, value="нет связи")
                sheet.cell(row=row[0].row, column=6, value="ошибка")
                print(f"IP {ip} недоступен, записано 'нет связи'.")
                no_connection_count += 1
            else:
                # Отправляем запрос к API для получения версии
                version = get_firmware(ip, protocol)

                # Записываем полученные данные в Excel
                sheet.cell(row=row[0].row, column=5, value=version)  # Версия (столбец E)
                
                # Записываем статус обработки
                # Проверяем, начинается ли версия с "Поток" или "ITC"
                if version and (str(version).startswith("Поток") or str(version).startswith("ITC")):
                    status = "успешно"
                    success_count += 1
                else:
                    status = "ошибка"
                    error_count += 1
                
                sheet.cell(row=row[0].row, column=6, value=status)  # Статус (столбец F)

                # Выводим результат в консоль
                print(f"IP: {ip}, Тип ДК: {protocol}, Версия: {version}, Статус: {status}")
        else:
            print(f"Пропущена строка {row[0].row} с неверными данные: IP = {ip}, Тип ДК = {protocol}")
            sheet.cell(row=row[0].row, column=5, value="неверные данные")
            sheet.cell(row=row[0].row, column=6, value="ошибка")
            error_count += 1

    # Сохраняем обновленный файл через диалоговое окно
    save_path = get_save_file()
    if save_path:
        wb.save(save_path)
        print(f"Файл сохранен как: {save_path}")
    else:
        print("Файл не сохранен.")

    # Выводим статистику
    print("\n" + "="*50)
    print("СТАТИСТИКА ОБРАБОТКИ:")
    print("="*50)
    print(f"Всего строк: {total_processed}")
    print(f"Успешно обработано: {success_count}")
    print(f"Перепроверено строк: {rechecked_count}")
    print(f"Пропущено (обработано ранее): {previously_processed_correct}")
    print(f"Нет связи: {no_connection_count}")
    print(f"Проверьте IP: {invalid_ip_count}")
    print(f"Ошибки (другие): {error_count}")
    print("="*50)

# Функция для запроса файла у пользователя
def get_file():
    # Создаем скрытое окно для выбора файла
    root = Tk()
    root.withdraw()  # Скрыть главное окно

    # Открываем диалог для выбора файла
    file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Excel files", "*.xlsx;*.xlsm")])

    if not file_path:  # Если файл не выбран
        print("Файл не выбран. Пожалуйста, выберите файл.")
        exit()

    return file_path

# Функция для запроса места для сохранения файла
def get_save_file():
    # Создаем скрытое окно для выбора места сохранения
    root = Tk()
    root.withdraw()  # Скрыть главное окно

    # Открываем диалог для выбора места сохранения файла
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    return save_path

if __name__ == "__main__":
    # Получаем файл от пользователя
    file_path = get_file()

    print(f"Выбран файл: {file_path}")
    
    # Запускаем синхронную обработку
    process_excel(file_path)
