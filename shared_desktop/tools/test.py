from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

def create_table(controller_type="Поток", num_klemm=12):
    """Создает таблицу с заданным типом контроллера и количеством клемм"""
    
    wb = Workbook()
    ws = wb.active
    
    if controller_type == "Поток":
        ws.title = "Таблица ТК Поток"
        create_table_potok(ws, num_klemm)
    else:  # Поток-Д
        ws.title = "Таблица ТК Поток-Д"
        create_table_potok_d(ws)
    
    return wb

def create_table_potok(ws, num_klemm):
    """Создает таблицу для контроллера Поток"""
    setup_worksheet(ws)
    create_header(ws, "Поток")
    create_table_headers(ws)
    create_klemmas_potok(ws, num_klemm)
    create_footer(ws, num_klemm)

def create_table_potok_d(ws):
    """Создает таблицу для контроллера Поток-Д (всегда 12 клемм)"""
    setup_worksheet(ws)
    create_header(ws, "Поток-Д")
    create_table_headers(ws)
    create_klemmas_potok_d(ws)
    create_footer_potok_d(ws)

def setup_worksheet(ws):
    """Настройка размеров столбцов и стилей"""
    column_widths = {
        'A': 6, 'B': 10, 'C': 10, 'D': 12, 'E': 12,
        'F': 20, 'G': 14, 'H': 14, 'I': 8, 'J': 8
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def create_header(ws, controller_type):
    """Создание заголовка таблицы"""

    ws.row_dimensions[3].height = 35
    ws.merge_cells('A1:J1')
    ws.merge_cells('A2:J2')
    
    header_data = [
        (1, "Объект: CO"),
        (2, f"Таблица коммутации контроллера {controller_type}")
    ]
    
    center = Alignment(horizontal='center', vertical='center')
    
    for row, text in header_data:
        cell = ws[f'A{row}']
        cell.value = text
        cell.alignment = center
        if row == 2:
            cell.font = Font(bold=True)

def create_table_headers(ws):
    """Создание заголовков таблицы"""

    merges = ['A3:C3', 'E3:E3', 'F3:F3', 'G3:G3', 'H3:H3', 'I3:J3']
    for merge in merges:
        ws.merge_cells(merge)

    headers = [
        ('A3', '↓ Клеммы ↓'), ('D3', '№ тир-ра'), ('E3', '№ напр.'), 
        ('F3', 'Тип напр.'), ('G3', 'Разреш. фазы'), ('H3', '№ светофора'), 
        ('I3', 'Сигнал')
    ]

    bold_italic = Font(bold=True, italic=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    border_bold = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    
    set_merged_cell_borders(ws, 'A3:j3', border_bold)

    for cell_ref, value in headers:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = bold_italic
        cell.alignment = center
        cell.border = border_bold

def set_merged_cell_borders(ws, merge_range, border):
    """Устанавливает границы для всех ячеек в объединенном диапазоне"""
    from openpyxl.worksheet.cell_range import CellRange
    cr = CellRange(merge_range)
    
    for row in range(cr.min_row, cr.max_row + 1):
        for col in range(cr.min_col, cr.max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border

def create_klemmas_potok(ws, num_klemm):
    """Создание клемм для Поток (4 строки на клемму)"""
    current_row = 4
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    
    for klemma_num in range(1, num_klemm + 1):
        create_klemma_potok(ws, current_row, klemma_num, bold_font, center, grey_fill)
        current_row += 4
        
        if (klemma_num == 2) or (klemma_num > 2 and (klemma_num - 2) % 4 == 0 and klemma_num != num_klemm):
            create_separator(ws, current_row, bold_font, center, blue_fill)
            current_row += 1

def create_klemma_potok(ws, start_row, klemma_num, bold_font, center, grey_fill):
    """Создание одной клеммы для Поток (4 строки)"""
    merges = [
        f'A{start_row}:A{start_row+3}', f'E{start_row}:E{start_row+3}',
        f'G{start_row}:G{start_row+3}', f'H{start_row}:H{start_row+3}',
        f'I{start_row}:J{start_row}', f'I{start_row+1}:J{start_row+1}',
        f'I{start_row+2}:J{start_row+2}', f'I{start_row+3}:J{start_row+3}'
    ]
    for merge in merges:
        ws.merge_cells(merge)

    klemma_data = [
        (start_row, 'A', klemma_num),
        (start_row, 'B', 'О'),
        (start_row, 'C', ' О'),
        (start_row, 'D', '1'),
        (start_row, 'I', 'красный'),
        (start_row+1, 'B', 'О'),
        (start_row+1, 'C', ' О'),
        (start_row+1, 'D', f'=D{start_row}+1'),
        (start_row+1, 'I', 'желтый'),
        (start_row+2, 'B', 'О'),
        (start_row+2, 'C', ' О'),
        (start_row+2, 'D', f'=D{start_row+1}+1'),
        (start_row+2, 'I', 'зеленый'),
        (start_row+3, 'B', 'О'),
        (start_row+3, 'C', ' О'),
        (start_row+3, 'D', '4'),
        (start_row+3, 'I', 'стрелка'),
    ]

    apply_klemma_styles(ws, klemma_data, start_row, bold_font, center, grey_fill, 4)
    
    border_middle_thin = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    set_merged_cell_borders(ws, f'A{start_row}:J{start_row+3}', border_middle_thin)
    set_merged_cell_borders(ws, f'A{start_row}:J{start_row}', 
                           Border(left=Side(style='medium'), right=Side(style='medium'), 
                                 top=Side(style='medium'), bottom=Side(style='thin')))
    set_merged_cell_borders(ws, f'A{start_row+3}:J{start_row+3}', 
                           Border(left=Side(style='medium'), right=Side(style='medium'), 
                                 top=Side(style='thin'), bottom=Side(style='medium')))

def create_klemmas_potok_d(ws):
    """Создание клемм для Поток-Д (3 строки на клемму, разделитель после каждой)"""
    current_row = 4
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    
    for klemma_num in range(1, 13):  # Всегда 12 клемм для Поток-Д
        create_klemma_potok_d(ws, current_row, klemma_num, bold_font, center, grey_fill)
        current_row += 3
        
        # Разделитель после каждой клеммы, кроме последней
        if klemma_num < 12:
            create_separator(ws, current_row, bold_font, center, blue_fill)
            current_row += 1

def create_klemma_potok_d(ws, start_row, klemma_num, bold_font, center, grey_fill):
    """Создание одной клеммы для Поток-Д (3 строки)"""
    merges = [
        f'A{start_row}:A{start_row+2}', f'E{start_row}:E{start_row+2}',
        f'G{start_row}:G{start_row+2}', f'H{start_row}:H{start_row+2}',
        f'I{start_row}:J{start_row}', f'I{start_row+1}:J{start_row+1}',
        f'I{start_row+2}:J{start_row+2}'
    ]
    for merge in merges:
        ws.merge_cells(merge)

    klemma_data = [
        (start_row, 'A', klemma_num),
        (start_row, 'B', 'О'),
        (start_row, 'C', ' О'),
        (start_row, 'D', '1'),
        (start_row, 'I', 'красный'),
        (start_row+1, 'B', 'О'),
        (start_row+1, 'C', ' О'),
        (start_row+1, 'D', f'=D{start_row}+1'),
        (start_row+1, 'I', 'желтый'),
        (start_row+2, 'B', 'О'),
        (start_row+2, 'C', ' О'),
        (start_row+2, 'D', f'=D{start_row+1}+1'),
        (start_row+2, 'I', 'зеленый'),
    ]

    apply_klemma_styles(ws, klemma_data, start_row, bold_font, center, grey_fill, 3)
    
    border_middle_thin = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    set_merged_cell_borders(ws, f'A{start_row}:J{start_row+2}', border_middle_thin)
    set_merged_cell_borders(ws, f'A{start_row}:J{start_row}', 
                           Border(left=Side(style='medium'), right=Side(style='medium'), 
                                 top=Side(style='medium'), bottom=Side(style='thin')))
    set_merged_cell_borders(ws, f'A{start_row+2}:J{start_row+2}', 
                           Border(left=Side(style='medium'), right=Side(style='medium'), 
                                 top=Side(style='thin'), bottom=Side(style='medium')))

def apply_klemma_styles(ws, klemma_data, start_row, bold_font, center, grey_fill, rows_per_klemma):
    """Применение стилей к ячейкам клеммы"""
    border_top_bold = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='medium'), bottom=Side(style='thin')
    )
    
    border_middle_thin = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    border_bottom_bold = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='thin'), bottom=Side(style='medium')
    )
    
    for row, col, value in klemma_data:
        cell = ws[f'{col}{row}']
        cell.value = value
        cell.alignment = center
        
        if row == start_row:
            if col in ['B', 'C', 'D']:
                cell.border = border_top_bold
        elif row == start_row + rows_per_klemma - 1:
            if col in ['B', 'C', 'D']:
                cell.border = border_bottom_bold
        else:
            if col in ['B', 'C', 'D']:
                cell.border = border_middle_thin
        
        if col in ['A', 'B', 'C']:
            cell.font = bold_font
        
        if col in ['B', 'C']:
            cell.fill = grey_fill

def create_separator(ws, row, bold_font, center, blue_fill):
    """Создание разделителя между клеммами"""
    ws.merge_cells(f'D{row}:H{row}')
    ws.merge_cells(f'I{row}:J{row}')

    separator_data = [
        (f'A{row}', '', blue_fill),
        (f'B{row}', 'О', blue_fill),
        (f'C{row}', ' О', blue_fill),
        (f'D{row}', '', None),  
        (f'I{row}', 'Общий', None),
    ]

    border_bold = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='medium'), bottom=Side(style='medium')
    )

    for cell_ref, value, fill in separator_data:
        cell = ws[cell_ref]
        cell.value = value
        cell.alignment = center
        cell.border = border_bold
        if fill:
            cell.fill = fill

    set_merged_cell_borders(ws, f'A{row}:J{row}', border_bold)

def create_footer(ws, num_klemm):
    """Создание завершающей части таблицы для Поток"""
    for row in range(7, 1000):
        if ws[f'B{row}'].value is None:
            footer_row = row
            break

    ws.merge_cells(f'A{footer_row}:A{footer_row+2}')
    ws.merge_cells(f'G{footer_row}:J{footer_row}')
    ws.merge_cells(f'G{footer_row+1}:J{footer_row+1}')
    ws.merge_cells(f'C{footer_row+2}:J{footer_row+2}')

    center = Alignment(horizontal='center', vertical='center')

    footer_data = [
        (footer_row, [
            ('B', 'О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
            ('C', ' О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
            ('E', '─', None, center),
            ('F', 'фаза', None, center),
            ('G', 'Вых.220в для подключения УЗСП, ТООВ и т.д.', None, Alignment(vertical='center'))
        ]),
        (footer_row+1, [
            ('B', 'О', PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), center),
            ('C', 'О', PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), center),
            ('E', '─', None, center),
            ('F', 'земля', None, center),
            ('G', 'Заземление', None, Alignment(vertical='center'))
        ])
    ]
    
    apply_footer_styles(ws, footer_data, footer_row)
    
    border_bold = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    
    set_merged_cell_borders(ws, f'A{footer_row}:J{footer_row+2}', border_bold)

def create_footer_potok_d(ws):
    """Создание завершающей части таблицы для Поток-Д"""
    for row in range(7, 1000):
        if ws[f'B{row}'].value is None:
            footer_row = row
            break

    ws.merge_cells(f'A{footer_row}:A{footer_row+2}')
    ws.merge_cells(f'G{footer_row}:J{footer_row}')
    ws.merge_cells(f'G{footer_row+1}:J{footer_row+1}')
    ws.merge_cells(f'C{footer_row+2}:J{footer_row+2}')

    center = Alignment(horizontal='center', vertical='center')

    footer_data = [
        (footer_row, [
            ('B', 'О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
            ('C', ' О', PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), center),
            ('E', '─', None, center),
            ('F', 'фаза', None, center),
            ('G', 'Вых.220в для подключения УЗСП, ТООВ и т.д.', None, Alignment(vertical='center'))
        ]),
        (footer_row+1, [
            ('B', 'О', PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"), center),
            ('C', 'О', PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), center),
            ('E', '─', None, center),
            ('F', 'земля', None, center),
            ('G', 'Заземление', None, Alignment(vertical='center'))
        ])
    ]
    
    apply_footer_styles(ws, footer_data, footer_row)
    
    border_bold = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    
    set_merged_cell_borders(ws, f'A{footer_row}:J{footer_row+2}', border_bold)

def apply_footer_styles(ws, footer_data, start_row):
    """Применение стилей к футеру"""
    bold_font = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    
    border_bold = Border(
        left=Side(style='medium'), right=Side(style='medium'), 
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    
    for row_offset, cells in footer_data:
        for col, value, fill, alignment in cells:
            cell = ws[f'{col}{row_offset}']
            cell.value = value
            cell.font = bold_font
            cell.alignment = alignment if alignment else center
            cell.border = border_bold
            
            if fill:
                cell.fill = fill

def get_user_input():
    """Получение параметров от пользователя"""
    root = tk.Tk()
    root.withdraw()
    
    # Выбор типа контроллера
    controller_type = simpledialog.askstring(
        "Тип контроллера", 
        "Выберите тип контроллера:\n1 - Поток\n2 - Поток-Д", 
        initialvalue="1"
    )
    
    if controller_type == "1":
        controller_type = "Поток"
        # Для Поток запрашиваем количество клемм
        num_klemm = simpledialog.askinteger(
            "Количество клемм", 
            "Введите количество клемм (для Поток):", 
            initialvalue=12, 
            minvalue=1, 
            maxvalue=128
        )
        if num_klemm is None:
            return None, None
    elif controller_type == "2":
        controller_type = "Поток-Д"
        num_klemm = 12  # Для Поток-Д всегда 12 клемм
    else:
        messagebox.showerror("Ошибка", "Неверный выбор типа контроллера")
        return None, None
    
    return controller_type, num_klemm

def save_file_dialog():
    """Открывает диалоговое окно для выбора места сохранения файла"""
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Сохранить таблицу ТК как...",
        initialfile="Таблица ТК.xlsx"
    )
    
    return file_path

def main():
    try:
        print("Выбор параметров таблицы...")
        controller_type, num_klemm = get_user_input()
        
        if controller_type is None:
            print("Создание таблицы отменено.")
            return
        
        print(f"Создание таблицы для контроллера {controller_type}...")
        wb = create_table(controller_type, num_klemm)
        
        save_path = save_file_dialog()
        if not save_path:
            print("Сохранение отменено.")
            return
        
        print("Сохранение файла...")
        wb.save(save_path)
        
        print(f"Файл '{save_path}' успешно создан.")
        
    except Exception as e:
        error_msg = f"Произошла ошибка: {str(e)}"
        messagebox.showerror("Ошибка", error_msg)
        print(error_msg)

if __name__ == "__main__":
    main()